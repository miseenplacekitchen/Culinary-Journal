#!/usr/bin/env python3
"""Probe Garden.xlsx structure for Phase 56 ingest mapping."""
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "brainstorm-inbox" / "2025.11.09_Garden.xlsx"
OUT = ROOT / "_excel_probe.txt"


def main() -> None:
    import openpyxl

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    lines = ["sheets: " + ", ".join(wb.sheetnames), ""]
    for name in [
        "Section Definitions",
        "Definitions",
        "(Quick) Plant Profile",
        "Plant Care Card Sheet",
        "Growth Calendar",
        "Master Sheet",
        "(Quick) Plant Care Card",
        "(Quick) Plant Care Calendar",
    ]:
        if name not in wb.sheetnames:
            lines.append(f"{name} MISSING\n")
            continue
        ws = wb[name]
        lines.append(f"=== {name} ===")
        n = 0
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [str(c).strip() if c is not None else "" for c in row[:8]]
            if any(vals):
                lines.append(f"{i}\t{vals}")
                n += 1
                if n >= 25:
                    break
        lines.append("")
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
