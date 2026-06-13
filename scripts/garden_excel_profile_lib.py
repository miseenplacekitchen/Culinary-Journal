"""Parse Garden.xlsx profile sheets → plants columns + care + calendar payloads."""
from __future__ import annotations

import re
from pathlib import Path

from garden_ingest_lib import esc_sql, slugify

XLSX_DEFAULT = Path(__file__).resolve().parents[1] / "brainstorm-inbox" / "2025.11.09_Garden.xlsx"

LOCATION_CLIMATE = {
    "brisbane": "humid-subtropical",
    "kerala": "tropical-monsoon",
    "thiruvalla": "tropical-monsoon",
}

# Inbox city labels → climate copy (never store city names in live profile text)
_CITY_COPY_REPLACEMENTS = (
    (re.compile(r"Brisbane's humid subtropical", re.I), "Humid subtropical"),
    (re.compile(r"Brisbane's", re.I), "Humid subtropical"),
    (re.compile(r"under Brisbane conditions", re.I), "in humid subtropical conditions"),
    (re.compile(r"\bin Brisbane\b", re.I), "in humid subtropical climates"),
    (re.compile(r"Brisbane summers", re.I), "humid subtropical summers"),
    (re.compile(r"Brisbane conditions", re.I), "humid subtropical conditions"),
    (re.compile(r"Brisbane:", re.I), "Humid subtropical:"),
    (re.compile(r"\bBrisbane\b", re.I), "humid subtropical climates"),
    (re.compile(r"Kerala's", re.I), "Tropical monsoon"),
    (re.compile(r"\bKerala\b", re.I), "tropical monsoon climates"),
)


def neutralize_location_copy(text: str) -> str:
    if not text:
        return text
    out = text
    for pattern, repl in _CITY_COPY_REPLACEMENTS:
        out = pattern.sub(repl, out)
    return out

# Excel cultivar/profile names → existing species shell slug
SPECIES_SLUG_ALIASES = {
    "black beauty tomato": "tomato",
    "purple romagna artichoke": "artichoke",
}

# Quick Plant Profile / Care Card titles → plants column
PROFILE_COLUMN_MAP = {
    "common plant name": "common_name",
    "plant name": "common_name",
    "botanical name": "botanical_name",
    "botanical name (binomial)": "botanical_name",
    "plant family": "plant_family",
    "plant type": "plant_type",
    "subspecies / taxonomic subgroup": "subspecies",
    "taxonomic authority": "taxonomic_authority",
    "genetic lineage type": "genetic_lineage_type",
    "variety / cultivar": "variety_cultivar",
    "origin": "origin",
    "growth habit": "growth_rate",
    "size at maturity": "size_height",
    "expected growth rate": "growth_rate",
    "pollination type": "pollination_type",
    "flowering season": "flowering_season",
    "propagation methods": "propagation_methods",
    "germination time": "germination_time",
    "time to harvest": "time_to_harvest",
    "seasonal planting windows": "planting_windows",
    "harvesting season": "harvest_season",
    "harvesting method": "harvesting_method",
    "care card summary": "care_summary",
}

# Section 3+ titles → plant_climate_care.field_key
CARE_FIELD_MAP = {
    "climate suitability": "climate",
    "soil management": "soil",
    "ph levels": "ph",
    "sunlight requirements": "sunlight",
    "wind resistance": "wind",
    "water management": "water",
    "frost tolerance": "frost",
    "seasonal risk indicators": "seasonal_risk",
    "fertilization schedule": "fertilisation",
    "mulching requirements": "mulching",
    "pruning requirements": "pruning",
    "special care notes": "special_care",
    "rotation category": "rotation",
    "pest management": "pest_mgmt",
    "disease pressure notes": "disease_notes",
    "water requirements": "water",
    "light requirements": "sunlight",
    "companion plant recommendations": "companions",
    "incompatible plants": "incompatibles",
}

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def normalize_title(title: str) -> str:
    return re.sub(r"\s+", " ", (title or "").strip().lower())


def parse_core_risk_fix(text: str) -> tuple[str, str, str]:
    if not text or text in ("#N/A", "-"):
        return "", "", ""
    core_parts: list[str] = []
    risk_parts: list[str] = []
    fix_parts: list[str] = []
    for raw in text.replace("\r", "").split("\n"):
        line = raw.strip().lstrip("•").strip()
        if not line:
            continue
        low = line.lower()
        if low.startswith("core:"):
            core_parts.append(line.split(":", 1)[1].strip())
        elif low.startswith("risk:"):
            risk_parts.append(line.split(":", 1)[1].strip())
        elif low.startswith("fix:"):
            fix_parts.append(line.split(":", 1)[1].strip())
        elif not core_parts and not risk_parts and not fix_parts:
            core_parts.append(line)
        else:
            core_parts.append(line)
    core = " ".join(core_parts).strip()
    risk = " ".join(risk_parts).strip()
    fix = " ".join(fix_parts).strip()
    if core and not risk and "risk" in core.lower():
        return core[:400], "", ""
    return core[:500], risk[:500], fix[:500]


def guess_species_slug(profile_label: str, plant_name: str = "") -> str:
    label = (profile_label or plant_name or "").lower().strip()
    label = re.sub(r"\s*-\s*(brisbane|kerala|thiruvalla)\s*$", "", label, flags=re.I).strip()
    if label in SPECIES_SLUG_ALIASES:
        return SPECIES_SLUG_ALIASES[label]
    for key, slug in SPECIES_SLUG_ALIASES.items():
        if key in label:
            return slug
    words = re.sub(r"[^a-z0-9\s]", " ", label).split()
    if words:
        return slugify(words[-1])
    return slugify(label)


def climate_for_location(location: str) -> str:
    loc = (location or "").lower()
    for key, slug in LOCATION_CLIMATE.items():
        if key in loc:
            return slug
    return "humid-subtropical"


def read_key_value_sheet(ws, start_row: int = 1) -> dict[str, str]:
    """Read rows with SNo + Titles + Description columns."""
    fields: dict[str, str] = {}
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        cells = list(row) + [None, None, None, None]
        title = cells[2]
        desc = cells[3]
        if title is None or str(title).strip() == "":
            continue
        title_s = str(title).strip()
        if title_s.startswith("🌿") or title_s.startswith("🌱") or title_s.startswith("🌤"):
            continue
        if str(cells[1] or "").strip().lower() == "sno.":
            continue
        val = "" if desc is None else str(desc).strip()
        if val in ("#N/A", "-", ""):
            continue
        fields[normalize_title(title_s)] = neutralize_location_copy(val)
    return fields


def parse_quick_plant_profile(ws) -> dict:
    profile_key = ""
    plant_name = ""
    location = ""
    for row in ws.iter_rows(max_row=6, values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if "please select the plant here:" in " ".join(cells).lower():
            for c in cells:
                if c and "please select" not in c.lower():
                    profile_key = c
        if cells[1].lower() == "plant name" and cells[3]:
            plant_name = cells[3]
        if plant_name and cells[3] and cells[3].lower() in ("brisbane", "kerala", "thiruvalla"):
            location = cells[3]
    fields = read_key_value_sheet(ws, start_row=6)
    return {
        "profile_key": profile_key,
        "plant_name": plant_name or profile_key.split(" - ")[0].strip(),
        "location": location or (profile_key.split(" - ")[-1].strip() if " - " in profile_key else "Brisbane"),
        "fields": fields,
    }


def profile_to_plant_updates(fields: dict[str, str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for title, val in fields.items():
        col = PROFILE_COLUMN_MAP.get(title)
        if col:
            out[col] = neutralize_location_copy(val)
    if "size at maturity" in fields and "size_height" not in out:
        out["size_height"] = fields["size at maturity"]
    summary = fields.get("care card summary")
    if summary and "care_summary" not in out:
        out["care_summary"] = summary
    return out


def profile_to_care_rows(fields: dict[str, str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for title, val in fields.items():
        key = CARE_FIELD_MAP.get(title)
        if not key:
            continue
        core, risk, fix = parse_core_risk_fix(val)
        if not core and not risk and not fix:
            core = neutralize_location_copy(val[:500])
        else:
            core, risk, fix = neutralize_location_copy(core), neutralize_location_copy(risk), neutralize_location_copy(fix)
        rows.append({"field_key": key, "core": core, "risk": risk, "fix": fix})
    return rows


def parse_calendar_from_profile(fields: dict[str, str]) -> list[dict]:
    """Derive coarse calendar rows from planting/harvest text when Growth Calendar grid unavailable."""
    out: list[dict] = []
    windows = fields.get("seasonal planting windows") or fields.get("propagation timing") or ""
    harvest = fields.get("harvesting season") or fields.get("flowering season") or ""
    if "august" in windows.lower() or "aug" in windows.lower():
        out.append({"activity": "sow", "month_start": 8, "month_end": 10, "notes": windows[:300]})
    if "november" in harvest.lower() or "nov" in harvest.lower():
        out.append({"activity": "harvest", "month_start": 11, "month_end": 3, "notes": harvest[:300]})
    if "transplant" in (fields.get("propagation transplanting") or "").lower():
        out.append({
            "activity": "transplant",
            "month_start": 9,
            "month_end": 11,
            "notes": fields.get("propagation transplanting", "")[:300],
        })
    return out


def list_master_profiles(wb) -> list[dict]:
    if "Master Sheet" not in wb.sheetnames:
        return []
    ws = wb["Master Sheet"]
    profiles: list[dict] = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if len(cells) < 7:
            continue
        validation = cells[5]
        location = cells[6]
        if not validation or validation in ("-", "Data Validation"):
            continue
        plant_id = cells[4] if len(cells) > 4 else ""
        profiles.append({
            "profile_key": validation,
            "plant_id": plant_id,
            "location": location or "Brisbane",
        })
    return profiles


def load_excel_profiles(xlsx_path: Path | None = None) -> list[dict]:
    import openpyxl

    path = xlsx_path or XLSX_DEFAULT
    if not path.exists():
        raise FileNotFoundError(f"Garden.xlsx not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    master = list_master_profiles(wb)
    if "(Quick) Plant Profile" not in wb.sheetnames:
        wb.close()
        return []

    # Workbook stores one active quick profile template; map master rows to it when labels match
    ws = wb["(Quick) Plant Profile"]
    quick = parse_quick_plant_profile(ws)
    wb.close()

    if not quick["fields"]:
        return []

    quick_species = guess_species_slug(quick["profile_key"] or quick["plant_name"])
    results: list[dict] = []
    seen: set[tuple[str, str]] = set()

    for m in master:
        key = m["profile_key"]
        slug = guess_species_slug(key)
        climate = climate_for_location(m["location"])
        dedupe = (slug, climate)
        if dedupe in seen:
            continue
        # Apply quick profile when same species or label matches active template
        if slug != quick_species and key.lower() not in quick["profile_key"].lower():
            continue
        seen.add(dedupe)
        fields = quick["fields"]
        results.append({
            "species_slug": slug,
            "profile_key": key,
            "location": m["location"],
            "climate_slug": climate,
            "plant_id_excel": m.get("plant_id", ""),
            "plant_updates": profile_to_plant_updates(fields),
            "care_rows": profile_to_care_rows(fields),
            "calendar_rows": parse_calendar_from_profile(fields),
        })
    return results


def sql_plant_block(entry: dict) -> list[str]:
    slug = entry["species_slug"]
    updates = entry["plant_updates"]
    if not updates:
        return [f"-- {entry['profile_key']} → {slug} (no plant column updates)", ""]
    set_parts = [f"{col} = '{esc_sql(val)}'" for col, val in updates.items()]
    set_parts.append("updated_at = now()")
    lines = [
        f"-- Excel: {entry['profile_key']} → species `{slug}` · climate `{entry['climate_slug']}`",
        "DO $$",
        "DECLARE v_plant uuid; v_cz uuid;",
        "BEGIN",
        f"  SELECT id INTO v_plant FROM public.plants WHERE slug = '{esc_sql(slug)}' LIMIT 1;",
        f"  IF v_plant IS NULL THEN RAISE NOTICE 'skip {esc_sql(slug)} — shell missing'; RETURN; END IF;",
        "  UPDATE public.plants SET " + ", ".join(set_parts) + " WHERE id = v_plant;",
        f"  SELECT id INTO v_cz FROM public.climate_zones WHERE slug = '{esc_sql(entry['climate_slug'])}' LIMIT 1;",
        "  IF v_cz IS NOT NULL THEN",
    ]
    for care in entry["care_rows"]:
        lines += [
            "    INSERT INTO public.plant_climate_care (plant_id, climate_zone_id, field_key, core, risk, fix)",
            f"    VALUES (v_plant, v_cz, '{esc_sql(care['field_key'])}', '{esc_sql(care['core'])}', '{esc_sql(care['risk'])}', '{esc_sql(care['fix'])}')",
            "    ON CONFLICT (plant_id, climate_zone_id, field_key) DO UPDATE SET",
            "      core = EXCLUDED.core, risk = EXCLUDED.risk, fix = EXCLUDED.fix;",
        ]
    for cal in entry["calendar_rows"]:
        lines += [
            "    INSERT INTO public.plant_calendar (plant_id, climate_zone_id, activity, month_start, month_end, notes)",
            f"    SELECT v_plant, v_cz, '{esc_sql(cal['activity'])}', {cal['month_start']}, {cal['month_end']}, '{esc_sql(cal['notes'])}'",
            "    WHERE NOT EXISTS (",
            "      SELECT 1 FROM public.plant_calendar pc",
            f"      WHERE pc.plant_id = v_plant AND pc.activity = '{esc_sql(cal['activity'])}'",
            f"        AND pc.month_start = {cal['month_start']} AND pc.month_end = {cal['month_end']}",
            "    );",
        ]
    lines += ["  END IF;", "END $$;", ""]
    return lines
